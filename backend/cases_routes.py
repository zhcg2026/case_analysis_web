import datetime
import io

import pandas as pd
from flask import jsonify, request, send_file
from sqlalchemy import func


def register_case_management_routes(
    app,
    Session,
    Case,
    CaseFollow,
    protected,
    get_json_payload,
    get_case_or_404,
    serialize_case,
    CASE_CATEGORIES,
    apply_case_category_fields,
    parse_pending_deadline,
):
    def _import_cases_from_excel(session, file):
        import tempfile
        import openpyxl
        import os
        import zipfile
        import uuid
        import xml.etree.ElementTree as ET

        upload_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads', 'case_photos')
        os.makedirs(upload_dir, exist_ok=True)

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp:
            file.save(temp.name)
            temp_path = temp.name

        try:
            file_size = os.path.getsize(temp_path)
            print(f"临时文件大小: {file_size} bytes")

            wb = openpyxl.load_workbook(temp_path)
            ws = wb.active
            image_map = {}

            try:
                with zipfile.ZipFile(temp_path, 'r') as zip_ref:
                    try:
                        with zip_ref.open('xl/drawings/_rels/drawing1.xml.rels') as f:
                            drawing_rels_xml = f.read()
                    except KeyError:
                        print("没有找到drawing关系文件")
                        drawing_rels_xml = None

                    if drawing_rels_xml:
                        drawing_rels_root = ET.fromstring(drawing_rels_xml)
                        rels_map = {}
                        for rel in drawing_rels_root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                            rel_id = rel.get('Id')
                            target = rel.get('Target')
                            if target.startswith('../media/'):
                                image_file = 'xl/media/' + target[9:]
                                rels_map[rel_id] = image_file

                        print(f"找到 {len(rels_map)} 个图片关系")

                        try:
                            with zip_ref.open('xl/drawings/drawing1.xml') as f:
                                drawing_xml = f.read()
                        except KeyError:
                            print("没有找到drawing文件")
                            drawing_xml = None

                        if drawing_xml:
                            drawing_root = ET.fromstring(drawing_xml)
                            for anchor in drawing_root.findall('.//{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}twoCellAnchor'):
                                from_elem = anchor.find('.//{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}from')
                                if from_elem is not None:
                                    row_elem = from_elem.find('.//{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}row')
                                    col_elem = from_elem.find('.//{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}col')
                                    if row_elem is not None and col_elem is not None:
                                        row = int(row_elem.text) + 1
                                        blip = anchor.find('.//{http://schemas.openxmlformats.org/drawingml/2006/main}blip')
                                        if blip is not None:
                                            embed = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                                            if embed and embed in rels_map:
                                                image_file = rels_map[embed]
                                                if row not in image_map:
                                                    image_map[row] = []
                                                image_map[row].append(image_file)

                            print(f"找到 {len(image_map)} 行包含图片")
            except Exception as e:
                print(f"提取图片信息失败: {str(e)}")
                import traceback
                traceback.print_exc()

            headers = [cell.value for cell in ws[1]]
            required_columns = ['任务号', '上报时间', '问题描述']
            for col in required_columns:
                if col not in headers:
                    return {'error': f'缺少必需列: {col}'}, 400

            task_number_col = headers.index('任务号')
            stage_light_col = headers.index('阶段红绿灯') if '阶段红绿灯' in headers else None
            auth_status_col = headers.index('阶段授权状态图标') if '阶段授权状态图标' in headers else None
            supervise_status_col = headers.index('阶段督办状态图标') if '阶段督办状态图标' in headers else None
            report_time_col = headers.index('上报时间') if '上报时间' in headers else None
            source_col = headers.index('问题来源') if '问题来源' in headers else None
            major_category_col = headers.index('大类名称') if '大类名称' in headers else None
            minor_category_col = headers.index('小类名称') if '小类名称' in headers else None
            problem_type_col = headers.index('问题类型') if '问题类型' in headers else None
            problem_desc_col = headers.index('问题描述') if '问题描述' in headers else None
            address_desc_col = headers.index('地址描述') if '地址描述' in headers else None
            responsible_grid_col = headers.index('责任网格') if '责任网格' in headers else None
            area_col = headers.index('所属区域') if '所属区域' in headers else None
            street_col = headers.index('所属街道') if '所属街道' in headers else None
            community_col = headers.index('所属社区') if '所属社区' in headers else None
            transfer_time_col = headers.index('批转时间') if '批转时间' in headers else None
            current_stage_time_info_col = headers.index('当前阶段时限信息') if '当前阶段时限信息' in headers else None
            current_stage_deadline_col = headers.index('当前阶段截止时间') if '当前阶段截止时间' in headers else None
            current_stage_remaining_time_col = headers.index('当前阶段剩余时间') if '当前阶段剩余时间' in headers else None
            area_level_col = headers.index('区域级别') if '区域级别' in headers else None
            area_level_name_col = headers.index('区域级别名称') if '区域级别名称' in headers else None
            responsible_area_name_col = headers.index('责属区域名称') if '责属区域名称' in headers else None
            bundle_deadline_col = headers.index('捆绑截止时间') if '捆绑截止时间' in headers else None
            bundle_time_limit_col = headers.index('捆绑截止时限') if '捆绑截止时限' in headers else None

            imported_count = 0
            skipped_count = 0
            for row_num in range(2, ws.max_row + 1):
                try:
                    task_number = ws.cell(row=row_num, column=task_number_col + 1).value
                    if not task_number:
                        skipped_count += 1
                        continue
                    task_number = str(task_number)

                    existing_case = session.query(Case).filter_by(task_number=task_number).first()
                    if existing_case:
                        skipped_count += 1
                        continue

                    photo_paths = []
                    if row_num in image_map:
                        images = image_map[row_num]
                        for img_idx, image_file in enumerate(images):
                            image_filename = f"case_{task_number}_{uuid.uuid4().hex}.jpeg"
                            image_path = os.path.join(upload_dir, image_filename)
                            try:
                                with zipfile.ZipFile(temp_path, 'r') as zip_ref:
                                    with zip_ref.open(image_file) as source:
                                        img_data = source.read()

                                if img_data:
                                    os.makedirs(os.path.dirname(image_path), exist_ok=True)
                                    with open(image_path, 'wb') as f:
                                        f.write(img_data)
                                    file_size = os.path.getsize(image_path)
                                    print(f"Image saved to {image_path}, size: {file_size} bytes")
                                    if file_size > 0:
                                        photo_paths.append(f"/uploads/case_photos/{image_filename}")
                                        print(f"Photo path added: /uploads/case_photos/{image_filename}")
                                    else:
                                        print(f"Warning: Image file is empty: {image_path}")
                                        if os.path.exists(image_path):
                                            os.remove(image_path)
                                else:
                                    print(f"No image data found for image {img_idx} at row {row_num}")
                            except Exception as img_error:
                                print(f"Error saving image {img_idx} for task {task_number}: {str(img_error)}")
                                import traceback
                                traceback.print_exc()
                                if os.path.exists(image_path) and os.path.getsize(image_path) == 0:
                                    os.remove(image_path)

                    photo_path = ','.join(photo_paths) if photo_paths else None
                    case_data = {
                        'task_number': task_number,
                        'stage_light': str(ws.cell(row=row_num, column=stage_light_col + 1).value) if stage_light_col is not None and ws.cell(row=row_num, column=stage_light_col + 1).value else None,
                        'auth_status': str(ws.cell(row=row_num, column=auth_status_col + 1).value) if auth_status_col is not None and ws.cell(row=row_num, column=auth_status_col + 1).value else None,
                        'supervise_status': str(ws.cell(row=row_num, column=supervise_status_col + 1).value) if supervise_status_col is not None and ws.cell(row=row_num, column=supervise_status_col + 1).value else None,
                        'report_time': ws.cell(row=row_num, column=report_time_col + 1).value if report_time_col is not None and ws.cell(row=row_num, column=report_time_col + 1).value else None,
                        'source': str(ws.cell(row=row_num, column=source_col + 1).value) if source_col is not None and ws.cell(row=row_num, column=source_col + 1).value else None,
                        'major_category': str(ws.cell(row=row_num, column=major_category_col + 1).value) if major_category_col is not None and ws.cell(row=row_num, column=major_category_col + 1).value else None,
                        'minor_category': str(ws.cell(row=row_num, column=minor_category_col + 1).value) if minor_category_col is not None and ws.cell(row=row_num, column=minor_category_col + 1).value else None,
                        'problem_type': str(ws.cell(row=row_num, column=problem_type_col + 1).value) if problem_type_col is not None and ws.cell(row=row_num, column=problem_type_col + 1).value else None,
                        'problem_desc': str(ws.cell(row=row_num, column=problem_desc_col + 1).value) if problem_desc_col is not None and ws.cell(row=row_num, column=problem_desc_col + 1).value else None,
                        'address_desc': str(ws.cell(row=row_num, column=address_desc_col + 1).value) if address_desc_col is not None and ws.cell(row=row_num, column=address_desc_col + 1).value else None,
                        'responsible_grid': str(ws.cell(row=row_num, column=responsible_grid_col + 1).value) if responsible_grid_col is not None and ws.cell(row=row_num, column=responsible_grid_col + 1).value else None,
                        'area': str(ws.cell(row=row_num, column=area_col + 1).value) if area_col is not None and ws.cell(row=row_num, column=area_col + 1).value else None,
                        'street': str(ws.cell(row=row_num, column=street_col + 1).value) if street_col is not None and ws.cell(row=row_num, column=street_col + 1).value else None,
                        'community': str(ws.cell(row=row_num, column=community_col + 1).value) if community_col is not None and ws.cell(row=row_num, column=community_col + 1).value else None,
                        'transfer_time': ws.cell(row=row_num, column=transfer_time_col + 1).value if transfer_time_col is not None and ws.cell(row=row_num, column=transfer_time_col + 1).value else None,
                        'current_stage_time_info': str(ws.cell(row=row_num, column=current_stage_time_info_col + 1).value) if current_stage_time_info_col is not None and ws.cell(row=row_num, column=current_stage_time_info_col + 1).value else None,
                        'current_stage_deadline': ws.cell(row=row_num, column=current_stage_deadline_col + 1).value if current_stage_deadline_col is not None and ws.cell(row=row_num, column=current_stage_deadline_col + 1).value else None,
                        'current_stage_remaining_time': str(ws.cell(row=row_num, column=current_stage_remaining_time_col + 1).value) if current_stage_remaining_time_col is not None and ws.cell(row=row_num, column=current_stage_remaining_time_col + 1).value else None,
                        'area_level': int(ws.cell(row=row_num, column=area_level_col + 1).value) if area_level_col is not None and ws.cell(row=row_num, column=area_level_col + 1).value else None,
                        'area_level_name': str(ws.cell(row=row_num, column=area_level_name_col + 1).value) if area_level_name_col is not None and ws.cell(row=row_num, column=area_level_name_col + 1).value else None,
                        'responsible_area_name': str(ws.cell(row=row_num, column=responsible_area_name_col + 1).value) if responsible_area_name_col is not None and ws.cell(row=row_num, column=responsible_area_name_col + 1).value else None,
                        'bundle_deadline': ws.cell(row=row_num, column=bundle_deadline_col + 1).value if bundle_deadline_col is not None and ws.cell(row=row_num, column=bundle_deadline_col + 1).value else None,
                        'bundle_time_limit': str(ws.cell(row=row_num, column=bundle_time_limit_col + 1).value) if bundle_time_limit_col is not None and ws.cell(row=row_num, column=bundle_time_limit_col + 1).value else None,
                        'photo_path': photo_path
                    }
                    new_case = Case(**case_data)
                    session.add(new_case)
                    imported_count += 1
                except Exception as e:
                    print(f"Error importing row {row_num}: {str(e)}")
                    skipped_count += 1
                    continue

            return {
                'message': '导入完成',
                'imported_count': imported_count,
                'skipped_count': skipped_count
            }, 200
        finally:
            if os.path.exists(temp_path):
                os.remove(temp_path)

    def _build_cases_export_file(session, category, status):
        query = session.query(Case)
        if category:
            query = query.filter(Case.category == category)
        if status:
            query = query.filter(Case.status == status)

        cases = query.all()
        data = []
        for c in cases:
            data.append({
                '任务号': c.task_number,
                '案件分类': c.category or '',
                '状态': c.status or '跟进中',
                '上报时间': c.report_time.strftime('%Y-%m-%d %H:%M') if c.report_time else '',
                '问题来源': c.source or '',
                '大类': c.major_category or '',
                '小类': c.minor_category or '',
                '问题描述': c.problem_desc or '',
                '地址': c.address_desc or '',
                '责属区域': c.responsible_area_name or '',
                '权属单位': c.owner_unit or '',
                '挂账原因': c.pending_reason or '',
                '预计处置时间': c.pending_deadline.strftime('%Y-%m-%d') if c.pending_deadline else '',
                '疑难类型': c.difficult_type or '',
                '跟进次数': c.follow_count or 0,
                '最近跟进': c.last_follow_time.strftime('%Y-%m-%d %H:%M') if c.last_follow_time else '',
                '结案时间': c.close_time.strftime('%Y-%m-%d %H:%M') if c.close_time else '',
                '结案说明': c.close_remark or '',
                '备注': c.remark or ''
            })

        df = pd.DataFrame(data)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='案件数据')
        output.seek(0)
        return output

    @app.route('/api/cases/import', methods=['POST'])
    @protected
    def import_cases():
        print("案件导入API被调用")
        session = Session()
        try:
            print("检查请求文件")
            if 'file' not in request.files:
                print("没有文件部分")
                return jsonify({'error': '缺少上传文件'}), 400

            file = request.files['file']
            print(f"获取文件: {file.filename}")
            if file.filename == '':
                print("没有选择文件")
                return jsonify({'error': '未选择文件'}), 400

            if not file.filename.endswith('.xlsx'):
                print("文件类型不正确")
                return jsonify({'error': '仅支持 xlsx 文件'}), 400

            print(f"文件大小: {file.content_length} bytes")

            result, status_code = _import_cases_from_excel(session, file)
            if status_code == 200:
                session.commit()
            return jsonify(result), status_code
        except Exception as e:
            session.rollback()
            print(f"Error in import_cases: {str(e)}")
            import traceback
            traceback.print_exc()
            return jsonify({'error': str(e)}), 500
        finally:
            session.close()

    @app.route('/api/cases', methods=['GET'])
    @protected
    def get_cases():
        session = Session()
        try:
            page = request.args.get('page', 1, type=int)
            per_page = request.args.get('per_page', 20, type=int)
            search = request.args.get('search', '')
            category = request.args.get('category', '')
            status = request.args.get('status', '')

            query = session.query(Case)

            if category:
                query = query.filter(Case.category == category)
            if status:
                query = query.filter(Case.status == status)
            if search:
                search_filter = f"%{search}%"
                query = query.filter(
                    (Case.task_number.like(search_filter)) |
                    (Case.problem_desc.like(search_filter)) |
                    (Case.address_desc.like(search_filter)) |
                    (Case.major_category.like(search_filter)) |
                    (Case.minor_category.like(search_filter))
                )

            total = query.count()
            cases = query.order_by(Case.created_at.desc()).offset((page - 1) * per_page).limit(per_page).all()
            cases_list = [serialize_case(case) for case in cases]

            session.commit()
            return jsonify({
                'cases': cases_list,
                'total': total,
                'page': page,
                'per_page': per_page,
                'total_pages': (total + per_page - 1) // per_page
            }), 200
        except Exception as e:
            session.rollback()
            print(f"Error in get_cases: {str(e)}")
            import traceback
            traceback.print_exc()
            return jsonify({'error': str(e)}), 500
        finally:
            session.close()

    @app.route('/api/cases/<int:case_id>', methods=['GET'])
    @protected
    def get_case_detail(case_id):
        session = Session()
        try:
            case = session.query(Case).filter_by(id=case_id).first()
            if not case:
                return jsonify({'error': '案件不存在'}), 404

            case_detail = serialize_case(case, include_updated_at=True)
            session.commit()
            return jsonify(case_detail), 200
        except Exception as e:
            session.rollback()
            print(f"Error in get_case_detail: {str(e)}")
            import traceback
            traceback.print_exc()
            return jsonify({'error': str(e)}), 500
        finally:
            session.close()

    @app.route('/api/cases/stats', methods=['GET'])
    @protected
    def get_cases_stats():
        """获取案件统计信息"""
        session = Session()
        try:
            stats = {
                'total': 0,
                'non_jurisdiction': 0,
                'pending': 0,
                'difficult': 0,
                'follow_up': 0,
                'closed': 0,
                'expiring_soon': 0,
            }

            stats['total'] = session.query(Case).count()
            category_stats = session.query(
                Case.category,
                func.count(Case.id)
            ).group_by(Case.category).all()

            for cat, count in category_stats:
                if cat == '非我局管辖':
                    stats['non_jurisdiction'] = count
                elif cat == '挂账案件':
                    stats['pending'] = count
                elif cat == '疑难案件':
                    stats['difficult'] = count

            status_stats = session.query(
                Case.status,
                func.count(Case.id)
            ).group_by(Case.status).all()

            for status, count in status_stats:
                if status == '跟进中' or status is None:
                    stats['follow_up'] += count
                elif status == '已结案':
                    stats['closed'] = count

            seven_days_later = datetime.datetime.now() + datetime.timedelta(days=7)
            stats['expiring_soon'] = session.query(Case).filter(
                Case.category == '挂账案件',
                Case.pending_deadline != None,
                Case.pending_deadline <= seven_days_later,
                Case.pending_deadline >= datetime.datetime.now(),
                Case.status != '已结案'
            ).count()

            session.commit()
            return jsonify(stats), 200
        except Exception as e:
            session.rollback()
            print(f"Error in get_cases_stats: {str(e)}")
            return jsonify({'error': str(e)}), 500
        finally:
            session.close()

    @app.route('/api/cases/<int:case_id>/category', methods=['PUT'])
    @protected
    def update_case_category(case_id):
        """更新案件分类"""
        session = Session()
        try:
            data = get_json_payload()
            category = data.get('category')

            if category not in CASE_CATEGORIES:
                return jsonify({'error': '无效的分类'}), 400

            case, error_resp = get_case_or_404(session, case_id)
            if error_resp:
                return error_resp

            case.category = category
            case.status = '跟进中'
            apply_case_category_fields(case, category, data)

            session.commit()
            return jsonify({'message': '分类更新成功'}), 200
        except Exception as e:
            session.rollback()
            print(f"Error in update_case_category: {str(e)}")
            return jsonify({'error': str(e)}), 500
        finally:
            session.close()

    @app.route('/api/cases/<int:case_id>/follow', methods=['POST'])
    @protected
    def add_case_follow(case_id):
        """添加跟进记录"""
        session = Session()
        try:
            if CaseFollow is None:
                return jsonify({'error': '跟进功能暂不可用，请先运行数据库迁移'}), 500

            data = get_json_payload()
            follow_type = data.get('follow_type', '其他')
            content = data.get('content', '')
            follow_user = data.get('follow_user', '')

            if not content:
                return jsonify({'error': '跟进内容不能为空'}), 400

            case, error_resp = get_case_or_404(session, case_id)
            if error_resp:
                return error_resp

            new_follow = CaseFollow(
                case_id=case_id,
                follow_type=follow_type,
                content=content,
                follow_user=follow_user,
                follow_time=datetime.datetime.now()
            )
            session.add(new_follow)

            case.last_follow_time = datetime.datetime.now()
            case.follow_count = (case.follow_count or 0) + 1

            session.commit()
            return jsonify({'message': '跟进记录添加成功', 'follow_id': new_follow.id}), 200
        except Exception as e:
            session.rollback()
            print(f"Error in add_case_follow: {str(e)}")
            return jsonify({'error': str(e)}), 500
        finally:
            session.close()

    @app.route('/api/cases/<int:case_id>/follows', methods=['GET'])
    @protected
    def get_case_follows(case_id):
        """获取案件的跟进记录"""
        session = Session()
        try:
            if CaseFollow is None:
                return jsonify({'follows': []}), 200

            follows = session.query(CaseFollow).filter_by(case_id=case_id).order_by(CaseFollow.follow_time.desc()).all()

            follows_list = []
            for f in follows:
                follows_list.append({
                    'id': f.id,
                    'follow_type': f.follow_type,
                    'content': f.content,
                    'attachments': f.attachments,
                    'follow_time': f.follow_time.strftime('%Y-%m-%d %H:%M') if f.follow_time else None,
                    'follow_user': f.follow_user
                })

            session.commit()
            return jsonify({'follows': follows_list}), 200
        except Exception as e:
            session.rollback()
            print(f"Error in get_case_follows: {str(e)}")
            return jsonify({'error': str(e)}), 500
        finally:
            session.close()

    @app.route('/api/cases/<int:case_id>/close', methods=['PUT'])
    @protected
    def close_case(case_id):
        """结案"""
        session = Session()
        try:
            data = get_json_payload()
            close_remark = data.get('close_remark', '')

            case, error_resp = get_case_or_404(session, case_id)
            if error_resp:
                return error_resp

            case.status = '已结案'
            case.close_time = datetime.datetime.now()
            case.close_remark = close_remark

            session.commit()
            return jsonify({'message': '结案成功'}), 200
        except Exception as e:
            session.rollback()
            print(f"Error in close_case: {str(e)}")
            return jsonify({'error': str(e)}), 500
        finally:
            session.close()

    @app.route('/api/cases/<int:case_id>', methods=['PUT'])
    @protected
    def update_case(case_id):
        """更新案件信息"""
        session = Session()
        try:
            data = get_json_payload()

            case, error_resp = get_case_or_404(session, case_id)
            if error_resp:
                return error_resp

            updatable_fields = [
                'owner_unit', 'contact_person', 'contact_phone',
                'pending_reason', 'pending_deadline', 'difficult_type', 'remark'
            ]

            for field in updatable_fields:
                if field in data:
                    if field == 'pending_deadline':
                        setattr(case, field, parse_pending_deadline(data[field]))
                    else:
                        setattr(case, field, data[field])

            session.commit()
            return jsonify({'message': '更新成功'}), 200
        except Exception as e:
            session.rollback()
            print(f"Error in update_case: {str(e)}")
            return jsonify({'error': str(e)}), 500
        finally:
            session.close()

    @app.route('/api/cases/<int:case_id>', methods=['DELETE'])
    @protected
    def delete_case(case_id):
        """删除案件"""
        session = Session()
        try:
            case, error_resp = get_case_or_404(session, case_id)
            if error_resp:
                return error_resp

            # 先删除关联的跟进记录
            if CaseFollow:
                session.query(CaseFollow).filter(CaseFollow.case_id == case_id).delete()

            # 删除案件
            session.delete(case)
            session.commit()
            return jsonify({'message': '删除成功'}), 200
        except Exception as e:
            session.rollback()
            print(f"Error in delete_case: {str(e)}")
            return jsonify({'error': str(e)}), 500
        finally:
            session.close()

    @app.route('/api/cases/export', methods=['GET'])
    @protected
    def export_cases():
        """导出案件数据"""
        session = Session()
        try:
            category = request.args.get('category', '')
            status = request.args.get('status', '')
            output = _build_cases_export_file(session, category, status)

            session.close()
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'案件导出_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )
        except Exception as e:
            session.rollback()
            print(f"Error in export_cases: {str(e)}")
            return jsonify({'error': str(e)}), 500
        finally:
            session.close()
